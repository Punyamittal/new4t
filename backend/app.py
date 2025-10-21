from flask import Flask, request, jsonify
from flask_cors import CORS
import pandas as pd
import os
from datetime import datetime
import json
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Hotel configuration
HOTEL_EXCEL_FILE_PATH = 'hotels.xlsx'
HOTEL_SHEET_NAME = 'Hotels'
HOTEL_HEADERS = ['Hotel Code', 'Name', 'Rating', 'Address', 'City ID', 'Country Code', 
                 'Latitude', 'Longitude', 'Facilities', 'Images', 'Created At']

# Room configuration
ROOM_EXCEL_FILE_PATH = 'hotel_rooms.xlsx'
ROOM_SHEET_NAME = 'Rooms'
ROOM_HEADERS = ['Room ID', 'Hotel Code', 'Booking Code', 'Room Name', 'Base Price', 
                'Total Fare', 'Currency', 'Is Refundable', 'Day Rates', 'Extras', 'Created At']

def create_hotel_excel_file_if_not_exists():
    """Create hotel Excel file with headers if it doesn't exist"""
    if not os.path.exists(HOTEL_EXCEL_FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = HOTEL_SHEET_NAME
        
        # Add headers
        for col, header in enumerate(HOTEL_HEADERS, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Save the workbook
        wb.save(HOTEL_EXCEL_FILE_PATH)
        print(f"Created new hotel Excel file: {HOTEL_EXCEL_FILE_PATH}")

def save_hotel_to_excel(data):
    """Save hotel data to Excel file using openpyxl"""
    try:
        # Create file if it doesn't exist
        create_hotel_excel_file_if_not_exists()
        
        # Load existing workbook
        wb = load_workbook(HOTEL_EXCEL_FILE_PATH)
        ws = wb[HOTEL_SHEET_NAME]
        
        # Find the next empty row
        next_row = ws.max_row + 1
        
        # Prepare data for the new row
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row_data = [
            data.get('hotel_code', ''),
            data.get('name', ''),
            data.get('rating', 0),
            data.get('address', ''),
            data.get('city_id', ''),
            data.get('country_code', ''),
            data.get('map_lat', 0),
            data.get('map_lon', 0),
            json.dumps(data.get('facilities', {})),
            json.dumps(data.get('images', [])),
            timestamp
        ]
        
        # Add data to the worksheet
        for col, value in enumerate(row_data, 1):
            ws.cell(row=next_row, column=col, value=value)
        
        # Save the workbook
        wb.save(HOTEL_EXCEL_FILE_PATH)
        print(f"Hotel data saved to Excel: {HOTEL_EXCEL_FILE_PATH}")
        return True
        
    except Exception as e:
        print(f"Error saving hotel data to Excel: {str(e)}")
        return False

def create_room_excel_file_if_not_exists():
    """Create room Excel file with headers if it doesn't exist"""
    if not os.path.exists(ROOM_EXCEL_FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = ROOM_SHEET_NAME
        
        # Add headers
        for col, header in enumerate(ROOM_HEADERS, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Save the workbook
        wb.save(ROOM_EXCEL_FILE_PATH)
        print(f"Created new room Excel file: {ROOM_EXCEL_FILE_PATH}")

def save_room_to_excel(data):
    """Save room data to Excel file using openpyxl"""
    try:
        # Create file if it doesn't exist
        create_room_excel_file_if_not_exists()
        
        # Load existing workbook
        wb = load_workbook(ROOM_EXCEL_FILE_PATH)
        ws = wb[ROOM_SHEET_NAME]
        
        # Find the next empty row
        next_row = ws.max_row + 1
        
        # Prepare data for the new row
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row_data = [
            data.get('room_id', ''),
            data.get('hotel_code', ''),
            data.get('booking_code', ''),
            data.get('room_name', ''),
            data.get('base_price', 0),
            data.get('total_fare', 0),
            data.get('currency', ''),
            data.get('is_refundable', False),
            json.dumps(data.get('day_rates', {})),
            json.dumps(data.get('extras', {})),
            timestamp
        ]
        
        # Add data to the worksheet
        for col, value in enumerate(row_data, 1):
            ws.cell(row=next_row, column=col, value=value)
        
        # Save the workbook
        wb.save(ROOM_EXCEL_FILE_PATH)
        print(f"Room data saved to Excel: {ROOM_EXCEL_FILE_PATH}")
        return True
        
    except Exception as e:
        print(f"Error saving room data to Excel: {str(e)}")
        return False

@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        "status": "success",
        "message": "Hotel booking backend is running"
    }), 200

@app.route('/hotel/add-hotel', methods=['POST'])
def add_hotel():
    """Handle hotel data submission and store in Excel"""
    try:
        # Get JSON data from request
        data = request.get_json()
        
        if not data:
            return jsonify({
                "success": False,
                "message": "No data provided"
            }), 400
        
        # Validate required fields
        required_fields = ['hotel_code', 'name', 'rating', 'address']
        missing_fields = [field for field in required_fields if not data.get(field)]
        
        if missing_fields:
            return jsonify({
                "success": False,
                "message": f"Missing required fields: {', '.join(missing_fields)}"
            }), 400
        
        # Save hotel data to Excel
        if save_hotel_to_excel(data):
            return jsonify({
                "success": True,
                "message": "Hotel added successfully"
            }), 200
        else:
            return jsonify({
                "success": False,
                "message": "Failed to save hotel data"
            }), 500
            
    except Exception as e:
        print(f"Error in add_hotel endpoint: {str(e)}")
        return jsonify({
            "success": False,
            "message": f"Server error while adding hotel"
        }), 500

@app.route('/hotelRoom/add', methods=['POST'])
def add_room():
    """Handle room data submission and store in Excel"""
    try:
        # Get JSON data from request
        data = request.get_json()
        
        if not data:
            return jsonify({
                "success": False,
                "message": "No data provided"
            }), 400
        
        # Validate required fields
        required_fields = ['room_id', 'hotel_code', 'booking_code', 'room_name']
        missing_fields = [field for field in required_fields if not data.get(field)]
        
        if missing_fields:
            return jsonify({
                "success": False,
                "message": f"Missing required fields: {', '.join(missing_fields)}"
            }), 400
        
        # Save room data to Excel
        if save_room_to_excel(data):
            return jsonify({
                "success": True,
                "message": "Hotel room added successfully"
            }), 200
        else:
            return jsonify({
                "success": False,
                "message": "Failed to save room data"
            }), 500
            
    except Exception as e:
        print(f"Error in add_room endpoint: {str(e)}")
        return jsonify({
            "success": False,
            "message": f"Server error while adding room"
        }), 500

@app.route('/hotels', methods=['GET'])
def get_hotels():
    """Get all hotels"""
    try:
        if not os.path.exists(HOTEL_EXCEL_FILE_PATH):
            return jsonify({
                "success": True,
                "data": [],
                "message": "No hotels found"
            }), 200
        
        df = pd.read_excel(HOTEL_EXCEL_FILE_PATH, sheet_name=HOTEL_SHEET_NAME)
        hotels = df.to_dict('records')
        
        return jsonify({
            "success": True,
            "data": hotels,
            "count": len(hotels)
        }), 200
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Error retrieving hotels: {str(e)}"
        }), 500

@app.route('/rooms', methods=['GET'])
def get_rooms():
    """Get all rooms"""
    try:
        if not os.path.exists(ROOM_EXCEL_FILE_PATH):
            return jsonify({
                "success": True,
                "data": [],
                "message": "No rooms found"
            }), 200
        
        df = pd.read_excel(ROOM_EXCEL_FILE_PATH, sheet_name=ROOM_SHEET_NAME)
        rooms = df.to_dict('records')
        
        return jsonify({
            "success": True,
            "data": rooms,
            "count": len(rooms)
        }), 200
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Error retrieving rooms: {str(e)}"
        }), 500

if __name__ == '__main__':
    # Create the Excel files if they don't exist when starting the server
    create_hotel_excel_file_if_not_exists()
    create_room_excel_file_if_not_exists()
    
    print("Starting Hotel Booking Backend Server...")
    print(f"Hotel Excel file: {os.path.abspath(HOTEL_EXCEL_FILE_PATH)}")
    print(f"Room Excel file: {os.path.abspath(ROOM_EXCEL_FILE_PATH)}")
    print("\nAvailable endpoints:")
    print("  POST /hotel/add-hotel - Add hotel details")
    print("  POST /hotelRoom/add - Add room details")
    print("  GET /hotels - Get all hotels")
    print("  GET /rooms - Get all rooms")
    print("  GET /health - Health check")
    
    app.run(debug=True, host='0.0.0.0', port=5000)
